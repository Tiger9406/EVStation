import PySimpleGUI as sg
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl

#vars for drawing figures in gui
_VARS = {'window': False,
         'fig_agg': False,
         'fig_agg2': False,
         'pltFig': False}

#allData gets the workbook file in project
allData = openpyxl.load_workbook("testingBook.xlsx")

#Home displays
#Left column
home_graph_column = [
    [sg.Canvas(key='figCanvas'), ]
]

#middle column with Title, intake file, intake month of file, and submit button
home_middle_column = [
    [sg.Text("EV Benefit"), ],
    [sg.FileBrowse("Enter new month", size=(20, 5), key='-in-'), sg.In(size=(10, 8), enable_events=True, key="time"), sg.Button("Submit")],
    [sg.Button("See Stats", size=(30, 8), key='gorange'), ],
]

#right column displaying total stats
home_right_column = [
    #separated two texts so easy to update second part
    [sg.Text("Total Stats", ), sg.Text("", key='-totalStats-'), ],
    [sg.Text("kWh:", ), sg.Text("", key='-totalKwh-'), ],
    [sg.Text("GHG:", ), sg.Text("", key='-totalGhg-'), ],
    [sg.Text("Sessions:", ), sg.Text("", key='-totalSessions-'), ],
    [sg.Text("Value of kWh:", ), sg.Text("", key='-totalValue-'), ],
    [sg.Text("Net Benefit:", ), sg.Text("", key='-totalBen-'), ],
    [sg.Text("Employee Paid: ", ), sg.Text("", key='-totalEmp-'), ],
]

#combining all displays into one
home_layout = [
    [
        sg.Column(home_graph_column),
        sg.VSeperator(),
        sg.Column(home_middle_column),
        sg.VSeperator(),
        sg.Column(home_right_column),
    ]
]

#Range page layouts; has a back button, and input boxes for user to enter range they want to see
range_layout = [
    [sg.Button("Back", key='gohome'),
     sg.Text("Stats")],
    [
        #  input months
        sg.Text("Range: "),
        sg.Input(key='range_time1', size=(5,3)),
        sg.Text("to "),
        sg.Input(key='range_time2', size=(5,3)),
        sg.Button("Go", key='gostats')
     ],
    [
        # Warning message
        sg.Text("Please enter proper months in proper format; ex: 01.21", key='properpls', visible=False)
    ],
]

#Stats page layout
stats_left = [
    [sg.Button("Back", key='gorange2'), sg.Button("Home", key='gohome2'), sg.Text("please just work")],
    [sg.Canvas(key='figCanvas2'), ],
    # can't directly get text of radio buttons so decided to add a hidden text with the same text
    # afterwards, so easy retrieval of x and y labels
    [sg.Text('x:'), sg.R('Month', "RADIO1", key='buttonxMonth', default=True),
     sg.R(text='kWh', group_id="RADIO1", key='buttonxD'), sg.Text('kWh', key="textxD", visible=False),
     sg.R(text='gHg (tCO2e)', group_id="RADIO1", key='buttonxE'),
     sg.Text('gHg (tCO2e)', key="textxE", visible=False),
     sg.R(text='Sessions (#)', group_id="RADIO1", key='buttonxC'),
     sg.Text('Sessions (#)', key="textxC", visible=False),
     sg.R(text='kWh Value ($)', group_id="RADIO1", key='buttonxG'),
     sg.Text('kWh Value ($)', key="textxG", visible=False),
     sg.R(text='Net Benefit ($)', group_id="RADIO1", key='buttonxI'),
     sg.Text('Net Benefit ($)', key="textxI", visible=False),
     sg.R(text='Employee Paid ($)', group_id="RADIO1", key='buttonxH'),
     sg.Text('Employee Paid ($)', key="textxH", visible=False), ],
    [sg.Text('y:'),
     sg.R(text='kWh', group_id="RADIO2", key='buttonyD', default=True),
     sg.Text('kWh', key="textyD", visible=False),
     sg.R(text='gHg (tCO2e)', group_id="RADIO2", key='buttonyE'),
     sg.Text('gHg (tCO2e)', key="textyE", visible=False),
     sg.R(text='Sessions (#)', group_id="RADIO2", key='buttonyC'),
     sg.Text('Sessions (#)', key="textyC", visible=False),
     sg.R(text='kWh Value ($)', group_id="RADIO2", key='buttonyG'),
     sg.Text('kWh Value ($)', key="textyG", visible=False),
     sg.R(text='Net Benefit ($)', group_id="RADIO2", key='buttonyI'),
     sg.Text('Net Benefit ($)', key="textyI", visible=False),
     sg.R(text='Employee Paid ($)', group_id="RADIO2", key='buttonyH'),
     sg.Text('Employee Paid ($)', key="textyH", visible=False),
     sg.Button("Submit", key='redoStats'),],
]
stats_right = [
    [sg.Text("Total Stats", ), sg.Text("", key='-statsStats-'), ],
    [sg.Text("kWh:", ), sg.Text("", key='-statsKwh-'), ],
    [sg.Text("GHG:", ), sg.Text("", key='-statsGhg-'), ],
    [sg.Text("Sessions:", ), sg.Text("", key='-statsSessions-'), ],
    [sg.Text("Value of kWh", ), sg.Text("", key='-statsValue-'), ],
    [sg.Text("Net Benefit:", ), sg.Text("", key='-statsBen-'), ],
    [sg.Text("Employee Paid: ", ), sg.Text("", key='-statsEmp-'), ],
]
stats_layout = [
    [
        sg.Column(stats_left),
        sg.VSeperator(),
        sg.Column(stats_right),
    ]
]

#overall layout; sets one visible and other two invisible; switch as maneuvers pages
layout = [
    [sg.Column(home_layout, key='homepage'), sg.Column(range_layout, visible=False, key='rangepage'),
           sg.Column(stats_layout, visible=False, key='statspage')],
          ]

# establish window
_VARS['window'] = sg.Window('EV Station', layout, finalize=True, resizable=False, location=(100, 100))

# returns a particular data in array format for each month, start to end
# ex [1, 6, 444, 4], 1 matching with the starting month and 4 corresponding to the last
def rangeData(start, end, category):
    array_data=[]
    counting = False
    for sheet in allData:
        if sheet.title == start:
            counting = True
        if counting:
            # in Excel, data starts at 2
            i = 2
            month_data = 0
            while sheet[f'A{i}'].value != None:
                if sheet[f'{category}{i}'].value != None:
                    month_data += sheet[f'{category}{i}'].value
                i += 1
            array_data.append(month_data)
        if sheet.title==end:
            break
    return array_data

# returns month in array format from start to end; works in partnership with rangeData()
def rangeMonth(start, end):
    array_month = []
    counting = False
    for sheet in allData:
        if sheet.title == start:
            counting = True
        if counting:
            array_month.append(sheet.title)
        if sheet.title == end:
            break
    return array_month

# drawing graph helper function; draws the plot on canvas
def draw_figure(canvas, figure):
    figure_canvas_agg = FigureCanvasTkAgg(figure, canvas)
    figure_canvas_agg.draw()
    figure_canvas_agg.get_tk_widget().pack(side='top', fill='both', expand=1)
    return figure_canvas_agg

# General draw Bar function, takes in custom things and puts custom graph on input canvas
def drawBar(categories, numbers, canvas, x_label, y_label, title_label):
    _VARS['pltFig'] = plt.figure(figsize = (7, 5))
    plt.bar(categories, numbers)
    plt.xlabel(x_label)
    plt.ylabel(y_label)
    plt.title(title_label)
    if canvas == 'figCanvas':
        _VARS['fig_agg'] = draw_figure(_VARS['window'][canvas].TKCanvas, _VARS['pltFig'])
    elif canvas == 'figCanvas2':
        if _VARS['fig_agg2'] != False:
            _VARS['fig_agg2'].get_tk_widget().forget()
        _VARS['fig_agg2'] = draw_figure(_VARS['window'][canvas].TKCanvas, _VARS['pltFig'])

# defScatter draws scatter plot with given inputs on given canvas
def drawScatter(xdata, ydata, canvas, x_label, y_label, title_label):
    # if pre-existing plot, erase
    if _VARS['fig_agg2'] != False:
        _VARS['fig_agg2'].get_tk_widget().forget()
    _VARS['pltFig'] = plt.figure(figsize=(6, 5))
    plt.scatter(xdata, ydata)
    plt.xlabel(x_label)
    plt.ylabel(y_label)
    plt.title(title_label)
    _VARS['fig_agg2'] = draw_figure(_VARS['window'][canvas].TKCanvas, _VARS['pltFig'])

# def drawMainBar basically draws the home graph: month (last five) vs net benefit
def drawMainBar():
    latestMonth = allData.sheetnames[-1]
    fifthLatestMonth = allData.sheetnames[-5]
    drawBar(rangeMonth(fifthLatestMonth, latestMonth), rangeData(fifthLatestMonth, latestMonth, 'I'),
            'figCanvas', 'Months', 'Net Benefit ($)', 'Net Benefit vs Time')

# def getRangeTotals returns an array of all numerical data over a certain range of months
# kwh, ghg, sessions, value kwh, net benefit, paid by employee
# I realize this is kind of mixed jumbo but this is in accordance with the order data is
# displayed on gui
def getRangeTotals(start, end):
    total = [0, 0, 0, 0, 0, 0]
    counting = False
    for sheet in allData:
        if sheet.title == start:
            counting = True
        if counting:
            i = 2
            # while there's an employee
            while sheet[f'A{i}'].value != None:
                # starts adding stuff
                if sheet[f'D{i}'].value != None:
                    total[0] += (sheet[f'D{i}'].value)
                if sheet[f'E{i}'].value != None:
                    total[1] += (sheet[f'E{i}'].value)
                if sheet[f'C{i}'].value != None:
                    total[2] += (sheet[f'C{i}'].value)
                if sheet[f'G{i}'].value != None:
                    total[3] += (sheet[f'G{i}'].value)
                if sheet[f'I{i}'].value != None:
                    total[4] += (sheet[f'I{i}'].value)
                if sheet[f'H{i}'].value != None:
                    total[5] += (sheet[f'H{i}'].value)
                i += 1
        if sheet.title == end:
            break
    return total

# updates summary stats on right side of home and stats page, depending on input
# btw this project made me love f strings
def updateStats(page, month1, month2):
    totalStats = getRangeTotals(month1, month2)
    _VARS['window'][f'-{page}Kwh-'].update(totalStats[0])
    _VARS['window'][f'-{page}Ghg-'].update(totalStats[1])
    _VARS['window'][f'-{page}Sessions-'].update(totalStats[2])
    _VARS['window'][f'-{page}Value-'].update(totalStats[3])
    _VARS['window'][f'-{page}Ben-'].update(totalStats[4])
    _VARS['window'][f'-{page}Emp-'].update(totalStats[5])

# check if the month inputted is in correct format
def ifDateValid(month):
    if month != None and len(month) == 5 and month[2] == '.':
        if month[0:2].isnumeric() and month[3:5].isnumeric():
            if int(month[0:2]) <= 12 and int(month[0:2]) > 0 and int(month[3:5])<=99:
                return True
    return False

# check if month is greater than the earliest month available
def ifDateGreater(month):
    if int(month[3:5]) > int(allData.sheetnames[0][3:5]):
        return True
    elif int(month[3:5]) == int(allData.sheetnames[0][3:5]):
        if int(month[0:2]) >= int(allData.sheetnames[0][0:2]):
            return True
    return False

# check if month is actually in the Excel book
def ifDateInRange(month):
    if int(month[3:5]) < int(allData.sheetnames[-1][3:5]) and int(month[3:5]) > int(allData.sheetnames[0][3:5]):
        return True
    elif int(month[3:5]) == int(allData.sheetnames[-1][3:5]) and int(month[0:2]) <= int(allData.sheetnames[-1][0:2]):
        return True
    elif int(month[3:5]) == int(allData.sheetnames[0][3:5]) and int(month[0:2]) >= int(allData.sheetnames[0][0:2]):
        return True
    return False

# find which radio button on stats page is selected, returns the key of the button
def findLabelButton(xy):
    if values[f'button{xy}D']:
        return f'button{xy}D'
    if values[f'button{xy}E']:
        return f'button{xy}E'
    if values[f'button{xy}C']:
        return f'button{xy}C'
    if values[f'button{xy}G']:
        return f'button{xy}G'
    if values[f'button{xy}I']:
        return f'button{xy}I'
    if values[f'button{xy}H']:
        return f'button{xy}H'
    return False

# presets when opening app
updateStats("total", allData.sheetnames[0], allData.sheetnames[-1])
drawMainBar()
layout = 1  # The currently visible layout
while True:
    event, values = _VARS['window'].read()
    if event in (None, 'Exit'):
        break
    elif event == 'gostats':
        # check if date entered in range page is valid, in format and if it's in the book
        if ifDateValid(values['range_time1']) and ifDateValid(values['range_time2']) and \
                ifDateInRange(values['range_time1']) and ifDateInRange(values['range_time2']):
            _VARS['window']['properpls'].update(visible=False)
            _VARS['window']['homepage'].update(visible=False)
            _VARS['window']['rangepage'].update(visible=False)
            # then it goes to stats page and draws a month vs net benefit bar graph by default
            _VARS['window']['statspage'].update(visible=True)
            drawBar(rangeMonth(values['range_time1'], values['range_time2']),
                    rangeData(values['range_time1'], values['range_time2'], 'I'),
                    'figCanvas2', 'Months', 'Net Benefit ($)', 'Months vs Net Benefit')
            updateStats("stats", values['range_time1'], values['range_time2'])
        else:
            # shows the warning signal
            _VARS['window']['properpls'].update(visible=True)
    # these two are just moving page events
    elif event == 'gorange' or event == 'gorange2':
        _VARS['window']['homepage'].update(visible=False)
        _VARS['window']['statspage'].update(visible=False)
        _VARS['window']['rangepage'].update(visible=True)
    elif event == 'gohome' or event == 'gohome2':
        _VARS['window']['rangepage'].update(visible=False)
        _VARS['window']['statspage'].update(visible=False)
        _VARS['window']['homepage'].update(visible=True)

        """
            What it does when the user submits is submits, takes the input workbook from submit button,
            makes a worksheet out of that workbook. It takes the sheetnames from all the data it has so 
            far, and gets the value from text entry about the month the data's about. It gets the 
            numerical value of the month and year.
            While the latest year is less than the year the client inputs, it just keeps adding sheets
            with increasing months and years. Once our data's latest sheet's name reaches the same year 
            as the input year, it starts increasing month until input month==latest. After that, there 
            will definitely exist a sheet with same month/year as the input, where the program will 
            copy all the data from the input data to the program's database's matching sheet and save 
            the new, edited workbook to a file inside the program."""
    # when user submits new data, first checks if it's proper format and it's greater than earliest
    elif event == 'Submit' and ifDateValid(values['time']) and ifDateGreater(values['time']):
        _VARS['fig_agg'].get_tk_widget().forget()  # erases main bar graph
        inputS = openpyxl.load_workbook(values['-in-'])  # gets the input workbook
        inputA = inputS.active  # gets the single sheet from input; can use this to get data
        allSheets = allData.sheetnames  # gets months in system
        time = values['time']  # the month user says data is from

        # this part if for if input month is later than the latest month, so creates empty
        # months until the inputted month
        lastMonth = int(allSheets[-1][0:2])
        lastYear = int(allSheets[-1][3:5])
        while int(time[3:5]) > lastYear:
            if lastMonth == 12:
                lastMonth = 1
                lastYear += 1
            else:
                lastMonth += 1
            if lastMonth < 10:
                allData.create_sheet("0" + str(lastMonth) + "." + str(lastYear))
            else:
                allData.create_sheet(str(lastMonth) + "." + str(lastYear))
        if int(time[3:5]) == lastYear:
            while int(time[0:2]) > lastMonth:
                lastMonth += 1
                if lastMonth < 10:
                    allData.create_sheet("0" + str(lastMonth) + "." + str(lastYear))
                else:
                    allData.create_sheet(str(lastMonth) + "." + str(lastYear))

        # finally done with creating sheets if it's too late
        # now move on to actually copying all data into book
        bookData = allData[time]
        # for data in inputA, copies into bookData (the input month in the system's book)
        for row in inputA.rows:
            for cell in row:
                bookData[cell.coordinate] = cell.value
        allData.save('testingBook.xlsx')
        # redraws main bar and updates stats on right of home page
        drawMainBar()
        updateStats("total", allData.sheetnames[0], allData.sheetnames[-1])
    elif event == 'redoStats':  # if user wants to change graph labels in stats page
        if values['buttonxMonth']:  # makes bar graph if x is month
            labelButton = findLabelButton('y')
            y_label = _VARS['window'][f'texty{labelButton[-1]}'].get()
            drawBar(rangeMonth(values['range_time1'], values['range_time2']),
                    rangeData(values['range_time1'], values['range_time2'], labelButton[-1]),
                    'figCanvas2', 'Months', y_label, 'Months vs '+ y_label)
        else:  # makes scatter graph otherwise
            labelButtonx = findLabelButton('x')
            x_label = _VARS['window'][f'textx{labelButtonx[-1]}'].get()
            labelButtony = findLabelButton('y')
            y_label = _VARS['window'][f'texty{labelButtony[-1]}'].get()
            drawScatter(rangeData(values['range_time1'], values['range_time2'], labelButtonx[-1]),
                        rangeData(values['range_time1'], values['range_time2'], labelButtony[-1]),
                        'figCanvas2', x_label, y_label, x_label + ' vs ' + y_label)

_VARS['window'].close()